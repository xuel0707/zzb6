if True:
    import os
    import json
    from openpyxl import load_workbook

    def extract_text_from_xlsx(xlsx_path):
        """Extract and merge all sheets from a .xlsx file into a single text string."""
        wb = load_workbook(xlsx_path, data_only=True)
        full_text_lines = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            full_text_lines.append(f"Sheet: {sheet_name}")  # 添加 sheet 标题

            for row in sheet.iter_rows(values_only=True):
                # 将每一行转换为以 \t 分隔的字符串
                line = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                full_text_lines.append(line)

            full_text_lines.append("")  # 不同 sheet 之间空一行

        return "\n".join(full_text_lines).strip()  # 合并为一个字符串


    def dir_to_json(directory):
        """Convert all .xlsx files in the given directory to a single JSON object with 'text' as the top-level key."""
        combined_text = []  # Initialize an empty list for collecting text from each file

        for filename in os.listdir(directory):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(directory, filename)
                print(f"Processing {file_path}")
                text = extract_text_from_xlsx(file_path)
                combined_text.append(text)  # Append the extracted text from each file

        # Combine all texts into one large string, separated by two newlines for clarity
        combined_text_str = "\n\n".join(combined_text).strip()

        return {'text': combined_text_str}  # Return as a dictionary with 'text' as the top-level key


    if __name__ == '__main__':
        import argparse
        parser = argparse.ArgumentParser(description='Convert all .xlsx files in a directory to a single JSON file with "text" as the top-level key.')
        parser.add_argument('directory', type=str, help='The path to the directory containing .xlsx files')
        parser.add_argument('output_json', type=str, help='The path where the output JSON file will be saved')

        args = parser.parse_args()
        combined_content = dir_to_json(args.directory)

        with open(args.output_json, 'w', encoding='utf-8') as json_file:
            json.dump(combined_content, json_file, ensure_ascii=False, indent=4)
else:
    
    import os
    import json
    from openpyxl import load_workbook

    def extract_sheet_from_xlsx(xlsx_path):
        """Extract each non-empty sheet from a .xlsx file and return them as separate text blocks."""
        wb = load_workbook(xlsx_path, data_only=True)
        sheets_text = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_lines = []
            has_content = False  # 标记是否是非空 sheet

            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else '' for cell in row]
                line = '\t'.join(row_values)
                sheet_lines.append(line)

                # 如果该行不是全空，则认为有内容
                if any(row_values):
                    has_content = True

            # 只保留有内容的 sheet
            if has_content:
                sheets_text.append({
                    "text": "\n".join(sheet_lines).strip()
                })

        return sheets_text


    def process_directory(directory):
        """Process all .xlsx files in the directory and collect each non-empty sheet into a list."""
        all_sheets_data = []

        for filename in os.listdir(directory):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(directory, filename)
                print(f"Processing {file_path}")
                sheets = extract_sheet_from_xlsx(file_path)
                all_sheets_data.extend(sheets)

        return all_sheets_data


    if __name__ == '__main__':
        import argparse

        parser = argparse.ArgumentParser(description='Convert each non-empty sheet of .xlsx files in a directory to JSON objects.')
        parser.add_argument('directory', type=str, help='Path to the directory containing .xlsx files')
        parser.add_argument('output_json', type=str, help='Path to save the output JSON file')

        args = parser.parse_args()

        sheets_data = process_directory(args.directory)

        with open(args.output_json, 'w', encoding='utf-8') as f:
            json.dump(sheets_data, f, ensure_ascii=False, indent=4)

        print(f"Total non-empty sheets extracted: {len(sheets_data)}")